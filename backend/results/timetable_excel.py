"""
timetable_excel.py — openpyxl Excel generator for timetables.

Layout mirrors the PDF format:
  - Rows  = days  (MON … SAT)
  - Cols  = time slots (period | SHORT BREAK | LUNCH BREAK)
  - Cell  = subject name only (or "LAB\n(name)" for labs)
  - Legend table below the grid: S.No | Subject Code | Name of Subject | Name of Faculty
  - Signature row at the bottom: Class Teacher | Head of Department | Principal
  - One sheet per section, plus a 'Faculty Load' summary sheet
"""

from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_points

# ---------------------------------------------------------------------------
# Colour palette (ARGB hex)
# ---------------------------------------------------------------------------
C_HEADER_BG = "FF1E3A5F"
C_HEADER_FG = "FFFFFFFF"
C_DAY_HDR_BG = "FF2D6A9F"
C_DAY_HDR_FG = "FFFFFFFF"
C_BREAK_BG = "FFF5E6C8"
C_LUNCH_BG = "FFFCE8B2"
C_LAB_BG = "FFFFF3CD"
C_ALT_ROW = "FFF5F5F5"
C_LEGEND_HDR = "FF37474F"
C_LEGEND_FG = "FFFFFFFF"
C_BORDER = "FFB0BEC5"
C_SIG_TOP = "FF555555"

DAYS_SHORT = ["MON", "TUE", "WED", "THU", "FRI", "SAT"]


def _fill(argb):
    return PatternFill(fill_type="solid", fgColor=argb)


def _border(color=C_BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _sig_border_top(color=C_SIG_TOP):
    """Border with only a top line (for signature cells)."""
    top = Side(style="medium", color=color)
    none = Side(style=None)
    return Border(top=top, left=none, right=none, bottom=none)


def _center(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


# ---------------------------------------------------------------------------
# Column-slot builder  (identical logic to timetable_pdf._build_column_slots)
# ---------------------------------------------------------------------------


def _build_column_slots(config):
    """Return ordered list of column descriptors.

    Each descriptor is a dict:
      { kind: "period"|"break"|"lunch", period_idx (period only), title }
    """
    start_time = config["start_time"]
    period_minutes = config["period_minutes"]
    periods_per_day = config["periods_per_day"]
    lunch_after = config["lunch_after_period"]
    lunch_minutes = config.get("lunch_minutes", 40)
    short_break_after = config.get("short_break_after_period")
    short_break_minutes = config.get("short_break_minutes", 10) or 10

    h, m = map(int, start_time.split(":"))
    cur = h * 60 + m

    def fmt(t):
        return "{:02d}:{:02d}".format((t // 60) % 24, t % 60)

    cols = []
    tidx = 0
    for p in range(periods_per_day):
        s, e = cur, cur + period_minutes
        cols.append(
            {
                "kind": "period",
                "period_idx": tidx,
                "title": "{}\n{}".format(fmt(s), fmt(e)),
            }
        )
        cur += period_minutes
        tidx += 1

        if short_break_after is not None and p == short_break_after:
            bs, be = cur, cur + short_break_minutes
            cols.append(
                {
                    "kind": "break",
                    "title": "SHORT\nBREAK\n{}-{}".format(fmt(bs), fmt(be)),
                }
            )
            cur += short_break_minutes

        if lunch_after is not None and p == lunch_after:
            ls, le = cur, cur + lunch_minutes
            cols.append(
                {
                    "kind": "lunch",
                    "title": "LUNCH\nBREAK\n{}-{}".format(fmt(ls), fmt(le)),
                }
            )
            cur += lunch_minutes

    return cols


# ---------------------------------------------------------------------------
# Section sheet writer
# ---------------------------------------------------------------------------


def _write_section_sheet(ws, config, solution, section, metadata, col_slots):
    working_days = config["working_days"]
    periods_per_day = config["periods_per_day"]
    college_name = metadata.get("college_name", "")
    branch = metadata.get("branch", "")
    year = metadata.get("year", "")
    semester = metadata.get("semester", "")
    room = metadata.get("room", "__________")

    total_cols = 1 + len(col_slots)  # DAY/TIME column + one column per slot

    # ── Row 1: College name ──────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws.cell(row=1, column=1, value=college_name)
    c.font = Font(name="Calibri", bold=True, size=14, color=C_HEADER_FG)
    c.fill = _fill(C_HEADER_BG)
    c.alignment = _center()
    ws.row_dimensions[1].height = 22

    # ── Row 2: Class / Room sub-header ───────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    sub_val = "Class: {} {} {} Sem \u2013 {} Section        Room Number: {}".format(
        year, branch, semester, section, room
    )
    c = ws.cell(row=2, column=1, value=sub_val)
    c.font = Font(name="Calibri", bold=True, size=10, color=C_HEADER_FG)
    c.fill = _fill(C_DAY_HDR_BG)
    c.alignment = _center()
    ws.row_dimensions[2].height = 18

    # ── Row 3: Column headers  (DAY/TIME | slot titles …) ───────────────────
    hdr_row = 3
    ws.row_dimensions[hdr_row].height = 42

    c = ws.cell(row=hdr_row, column=1, value="DAY/TIME")
    c.font = Font(bold=True, size=9, color=C_DAY_HDR_FG)
    c.fill = _fill(C_DAY_HDR_BG)
    c.alignment = _center()
    c.border = _border()

    for ci, cs in enumerate(col_slots, start=2):
        c = ws.cell(row=hdr_row, column=ci, value=cs["title"])
        c.alignment = _center()
        c.border = _border()
        if cs["kind"] == "break":
            c.font = Font(bold=True, size=7, color="FF7B5800")
            c.fill = _fill(C_BREAK_BG)
        elif cs["kind"] == "lunch":
            c.font = Font(bold=True, size=7, color="FF7B5800")
            c.fill = _fill(C_LUNCH_BG)
        else:
            c.font = Font(bold=True, size=8, color=C_DAY_HDR_FG)
            c.fill = _fill(C_DAY_HDR_BG)

    # ── Rows 4…: Day rows ────────────────────────────────────────────────────
    for d in range(working_days):
        data_row = hdr_row + 1 + d
        ws.row_dimensions[data_row].height = 30
        fill_bg = _fill(C_ALT_ROW) if d % 2 == 1 else _fill("FFFFFFFF")

        # Day label
        c = ws.cell(row=data_row, column=1, value=DAYS_SHORT[d])
        c.font = Font(bold=True, size=9, color=C_DAY_HDR_FG)
        c.fill = _fill(C_DAY_HDR_BG)
        c.alignment = _center()
        c.border = _border()

        for ci, cs in enumerate(col_slots, start=2):
            cell = ws.cell(row=data_row, column=ci)
            cell.alignment = _center()
            cell.border = _border()

            if cs["kind"] in ("break", "lunch"):
                cell.value = ""
                cell.fill = _fill(C_BREAK_BG if cs["kind"] == "break" else C_LUNCH_BG)
                continue

            p = cs["period_idx"]
            si = solution.get((d, p))

            if si is None:
                cell.value = ""
                cell.fill = fill_bg
            elif si.get("is_lab_cont"):
                cell.value = "(cont.)"
                cell.font = Font(italic=True, size=8, color="FF888888")
                cell.fill = _fill(C_LAB_BG)
            else:
                stype = si.get("type", "theory")
                sname = si.get("subject_name", "")
                cell.value = "LAB\n({})".format(sname) if stype == "lab" else sname
                cell.font = Font(bold=True, size=9) if stype != "lab" else Font(size=8)
                cell.fill = _fill(C_LAB_BG) if stype == "lab" else fill_bg

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 9  # DAY/TIME
    for ci, cs in enumerate(col_slots, start=2):
        col_letter = get_column_letter(ci)
        if cs["kind"] == "break":
            ws.column_dimensions[col_letter].width = 9
        elif cs["kind"] == "lunch":
            ws.column_dimensions[col_letter].width = 11
        else:
            ws.column_dimensions[col_letter].width = 14

    # ── Spacer row ───────────────────────────────────────────────────────────
    legend_start = hdr_row + working_days + 2  # one blank row gap

    # ── Legend header ────────────────────────────────────────────────────────
    leg_headers = ["S.No", "Subject Code", "Name of Subject", "Name of Faculty"]
    leg_widths = [6, 16, 36, 36]
    for ci, (hdr, w) in enumerate(zip(leg_headers, leg_widths), start=1):
        c = ws.cell(row=legend_start, column=ci, value=hdr)
        c.font = Font(bold=True, size=9, color=C_LEGEND_FG)
        c.fill = _fill(C_LEGEND_HDR)
        c.alignment = _center()
        c.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Legend data ──────────────────────────────────────────────────────────
    seen = {}
    sno = 1
    leg_row = legend_start + 1
    for p in range(periods_per_day):
        for d in range(working_days):
            si = solution.get((d, p))
            if si and not si.get("is_lab_cont"):
                code = si["subject_code"]
                if code not in seen:
                    seen[code] = True
                    fac = si["faculty_1"]
                    f2 = si.get("faculty_2", "")
                    if f2 and str(f2).lower() not in ("nan", "none", ""):
                        fac += ", {}".format(f2)
                    row_fill = _fill(C_ALT_ROW) if sno % 2 == 0 else _fill("FFFFFFFF")
                    for ci, val in enumerate(
                        [sno, code, si["subject_name"], fac], start=1
                    ):
                        c = ws.cell(row=leg_row, column=ci, value=val)
                        c.fill = row_fill
                        c.alignment = _center() if ci <= 2 else _left()
                        c.border = _border()
                        c.font = Font(size=9)
                    sno += 1
                    leg_row += 1

    # ── Signature row ────────────────────────────────────────────────────────
    sig_row = leg_row + 2
    ws.row_dimensions[sig_row].height = 36
    sig_labels = ["Class Teacher", "Head of the Department", "Principal"]
    for ci, label in enumerate(sig_labels, start=1):
        c = ws.cell(row=sig_row, column=ci, value=label)
        c.font = Font(bold=True, size=10, color=C_HEADER_BG)
        c.alignment = _center()
        c.border = _sig_border_top()

    # ── Footer ───────────────────────────────────────────────────────────────
    footer_row = sig_row + 1
    ws.merge_cells(
        start_row=footer_row, start_column=1, end_row=footer_row, end_column=total_cols
    )
    c = ws.cell(
        row=footer_row,
        column=1,
        value="Generated on {}".format(datetime.now().strftime("%d %b %Y, %I:%M %p")),
    )
    c.font = Font(size=7, color="FF888888")
    c.alignment = _center()


# ---------------------------------------------------------------------------
# Faculty load summary sheet
# ---------------------------------------------------------------------------


def _write_faculty_load_sheet(wb, full_solution, metadata):
    ws = wb.create_sheet(title="Faculty Load")

    headers = ["Faculty", "Section", "Subject Code", "Subject Name", "Weekly Hours"]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(
        row=1,
        column=1,
        value="Faculty Load Summary — {}".format(metadata.get("college_name", "")),
    )
    c.font = Font(bold=True, size=13, color=C_HEADER_FG)
    c.fill = _fill(C_HEADER_BG)
    c.alignment = _center()

    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = Font(bold=True, color=C_LEGEND_FG)
        cell.fill = _fill(C_LEGEND_HDR)
        cell.alignment = _center()
        cell.border = _border()

    row = 3
    for section, solution in sorted(full_solution.items()):
        faculty_hours = {}
        for (d, p), si in solution.items():
            if si.get("is_lab_cont"):
                continue
            for fk in ("faculty_1", "faculty_2"):
                fac = si.get(fk, "")
                if not fac or str(fac).lower() in ("nan", "none", ""):
                    continue
                key = (fac, section, si["subject_code"])
                if key not in faculty_hours:
                    faculty_hours[key] = {
                        "subject_name": si["subject_name"],
                        "count": 0,
                    }
                faculty_hours[key]["count"] += 1

        for (fac, sec, code), info in sorted(faculty_hours.items()):
            fill_c = C_ALT_ROW if row % 2 == 0 else "FFFFFFFF"
            for ci, val in enumerate(
                [fac, sec, code, info["subject_name"], info["count"]], start=1
            ):
                c = ws.cell(row=row, column=ci, value=val)
                c.fill = _fill(fill_c)
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border = _border()
            row += 1

    col_widths = [30, 10, 16, 30, 14]
    for ci, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def generate_timetable_excel(config, full_solution, metadata):
    """Generate an Excel workbook for all sections.

    Args:
        config        — schedule config dict
        full_solution — {section: {(day, period): slot_info}}
        metadata      — {college_name, branch, year, semester}

    Returns: BytesIO of the .xlsx file
    """
    wb = Workbook()
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    col_slots = _build_column_slots(config)

    for section, solution in sorted(full_solution.items()):
        ws = wb.create_sheet(title="Sec {}".format(section))
        _write_section_sheet(ws, config, solution, section, metadata, col_slots)

    _write_faculty_load_sheet(wb, full_solution, metadata)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
