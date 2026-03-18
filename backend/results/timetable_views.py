"""
timetable_views.py -- Django views for the timetable generator feature.

Endpoints:
  POST /api/timetable/generate/      -- receive config + Excel, solve, return JSON solution + AI summary
  POST /api/timetable/download/      -- download PDF or Excel file
  GET  /api/timetable/template/      -- download sample Excel input template
"""

import io
import os
import requests as http_requests

import pandas as pd
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from .timetable_solver import solve_timetable
from .timetable_pdf import (generate_timetable_pdf, generate_all_timetables_pdf,
                             generate_faculty_timetable_pdf, generate_all_faculty_timetables_pdf)
from .timetable_excel import generate_timetable_excel


OLLAMA_URL = os.environ.get("OLLAMA_URL", "http://spmvv_ollama:11434")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _parse_optional_int(val):
    if val is None:
        return None
    try:
        import math

        if isinstance(val, float) and math.isnan(val):
            return None
    except Exception:
        pass
    s = str(val).strip()
    if s == "" or s.lower() == "nan":
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def _parse_optional_str(val):
    if val is None:
        return None
    try:
        import math

        if isinstance(val, float) and math.isnan(val):
            return None
    except Exception:
        pass
    s = str(val).strip().lower()
    if s == "" or s == "nan":
        return None
    return s


def _parse_excel(file_obj):
    """Parse the uploaded Excel into a list of subject dicts.

    Required columns: subject_code, subject_name, type, hours_per_week,
                      faculty_1, branch, year, semester, section
    Optional columns: faculty_2, lab_consecutive_periods, lab_split_mode,
                      min_faculty_gap
    """
    try:
        df = pd.read_excel(file_obj)
    except Exception as e:
        raise ValueError("Could not read Excel file: {}".format(e))

    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]

    required_cols = {
        "subject_code",
        "subject_name",
        "type",
        "hours_per_week",
        "faculty_1",
        "branch",
        "year",
        "semester",
        "section",
    }
    missing = required_cols - set(df.columns)
    if missing:
        raise ValueError(
            "Missing columns in Excel: {}".format(", ".join(sorted(missing)))
        )

    subjects = []
    for _, row in df.iterrows():
        subj = {
            "subject_code": str(row["subject_code"]).strip(),
            "subject_name": str(row["subject_name"]).strip(),
            "type": str(row.get("type", "theory")).strip().lower(),
            "hours_per_week": int(row["hours_per_week"]),
            "faculty_1": str(row["faculty_1"]).strip(),
            "faculty_2": str(row.get("faculty_2", "") or "").strip(),
            "branch": str(row["branch"]).strip(),
            "year": str(row["year"]).strip(),
            "semester": str(row["semester"]).strip(),
            "section": str(row["section"]).strip(),
            "lab_consecutive_periods": _parse_optional_int(
                row.get("lab_consecutive_periods")
            ),
            "lab_split_mode": _parse_optional_str(row.get("lab_split_mode")),
            "min_faculty_gap": _parse_optional_int(row.get("min_faculty_gap")),
        }
        if subj["type"] not in ("theory", "lab"):
            subj["type"] = "theory"
        if subj["lab_split_mode"] is not None and subj["lab_split_mode"] not in (
            "split",
            "whole",
        ):
            subj["lab_split_mode"] = None
        subjects.append(subj)
    return subjects


def _derive_groups(subjects):
    """
    Derive the list of (branch, year, semester, [sections]) groups
    from the parsed subject list. section='ALL' means all sections
    found for that branch+year+semester combo.

    Returns a list of group dicts:
      { branch, year, semester, sections: [str, ...] }
    """
    from collections import defaultdict

    group_sections = defaultdict(set)
    group_has_all = defaultdict(bool)

    for s in subjects:
        key = (s["branch"], s["year"], s["semester"])
        sec = s["section"].upper()
        if sec == "ALL":
            group_has_all[key] = True
        else:
            group_sections[key].add(s["section"])

    all_keys = set(group_sections.keys()) | set(group_has_all.keys())
    groups = []
    for key in sorted(all_keys):
        branch, year, semester = key
        secs = sorted(group_sections[key]) if group_sections[key] else ["A"]
        groups.append(
            {
                "branch": branch,
                "year": year,
                "semester": semester,
                "sections": secs,
            }
        )
    return groups


def _build_config(data):
    """Validate and build global schedule config from POST form data.
    Branch/year/semester/sections are NOT required here -- they come from Excel.
    """

    def req(key, cast=str):
        val = data.get(key)
        if val is None or val == "":
            raise ValueError("Missing required field: {}".format(key))
        return cast(val)

    def opt(key, cast=str, default=None):
        val = data.get(key)
        if val is None or val == "":
            return default
        return cast(val)

    config = {
        "working_days": req("working_days", int),
        "periods_per_day": req("periods_per_day", int),
        "start_time": req("start_time"),
        "period_minutes": req("period_minutes", int),
        "lunch_after_period": req("lunch_after_period", int),
        "lunch_minutes": req("lunch_minutes", int),
        "short_break_after_period": opt("short_break_after_period", int),
        "short_break_minutes": opt("short_break_minutes", int, 10),
        "lab_consecutive_periods": req("lab_consecutive_periods", int),
        "min_faculty_gap": opt("min_faculty_gap", int, 1),
    }

    if config["working_days"] < 1 or config["working_days"] > 6:
        raise ValueError("working_days must be between 1 and 6.")
    if config["periods_per_day"] < 2 or config["periods_per_day"] > 12:
        raise ValueError("periods_per_day must be between 2 and 12.")

    return config


def _solution_to_serializable(solution):
    out = {}
    for section, slot_map in solution.items():
        out[section] = {}
        for (d, p), info in slot_map.items():
            out[section]["{},{}".format(d, p)] = info
    return out


def _solution_from_serializable(data):
    out = {}
    for section, slot_map in data.items():
        out[section] = {}
        for key, info in slot_map.items():
            d, p = map(int, key.split(","))
            out[section][(d, p)] = info
    return out


def _call_ollama(prompt):
    try:
        resp = http_requests.post(
            "{}/api/generate".format(OLLAMA_URL),
            json={
                "model": "qwen2.5:3b",
                "prompt": prompt,
                "stream": False,
                "options": {"temperature": 0.3, "num_predict": 400},
            },
            timeout=60,
        )
        if resp.status_code == 200:
            return resp.json().get("response", "").strip()
    except Exception:
        pass
    return ""


def _ai_summary(config, solution, groups, college_name):
    sections = list(solution.keys())
    total_slots = sum(
        1
        for sec in solution.values()
        for info in sec.values()
        if not info.get("is_lab_cont")
    )
    group_strs = ", ".join(
        "{} Year {} Sem {} ({})".format(
            g["branch"], g["year"], g["semester"], "/".join(g["sections"])
        )
        for g in groups
    )
    prompt = (
        "A university timetable has been generated for {college}. "
        "Groups scheduled: {groups}. "
        "Sections: {sections}. "
        "Working days: {days}, Periods per day: {ppd}. "
        "Total scheduled slots (excluding lunch/breaks): {total}. "
        "Default lab sessions are {lab_len} consecutive periods. "
        "Minimum faculty gap between classes: {gap} period(s). "
        "In 3-4 sentences, summarise the key characteristics of this timetable -- "
        "balance of theory vs lab, faculty load distribution, and any notable scheduling patterns. "
        "Be concise and professional."
    ).format(
        college=college_name,
        groups=group_strs,
        sections=", ".join(sections),
        days=config["working_days"],
        ppd=config["periods_per_day"],
        total=total_slots,
        lab_len=config["lab_consecutive_periods"],
        gap=config.get("min_faculty_gap", 1),
    )
    return _call_ollama(prompt)


def _ai_explain_failure(error_msg, config):
    prompt = (
        "A university timetable generator failed to find a valid schedule. "
        "The constraint solver reported:\n\n{error}\n\n"
        "Configuration used: {days} working days, "
        "{ppd} periods/day, "
        "lab blocks of {lab_len} consecutive periods, "
        "min faculty gap = {gap}. "
        "In 3-4 sentences, explain the likely cause in plain English and "
        "suggest 2-3 specific changes to the configuration or subject list "
        "that would make the problem solvable."
    ).format(
        error=error_msg,
        days=config["working_days"],
        ppd=config["periods_per_day"],
        lab_len=config["lab_consecutive_periods"],
        gap=config.get("min_faculty_gap", 1),
    )
    return _call_ollama(prompt)




def _build_faculty_solution(full_solution):
    """
    Invert a section-keyed solution into a faculty-keyed solution.

    Input:
      full_solution: { section: { (day, period): slot_info } }

    Output:
      { faculty_name: { (day, period): {
            section, subject_code, subject_name, type,
            is_lab_cont, lab_block_start, faculty_2
        } } }
    """
    from collections import defaultdict

    faculty_sol = defaultdict(dict)
    for section, slot_map in full_solution.items():
        for (d, p), info in slot_map.items():
            f1 = info.get("faculty_1", "")
            f2 = info.get("faculty_2", "")
            entry = {
                "section":       section,
                "subject_code":  info.get("subject_code", ""),
                "subject_name":  info.get("subject_name", ""),
                "type":          info.get("type", "theory"),
                "is_lab_cont":   info.get("is_lab_cont", False),
                "lab_block_start": info.get("lab_block_start", False),
                "faculty_2":     f2,
            }
            if f1 and str(f1).strip():
                faculty_sol[f1.strip()][(d, p)] = entry
            # For split labs faculty_2 is also teaching at the same slot
            if f2 and str(f2).strip() and str(f2).lower() not in ("nan", "none", ""):
                entry2 = {**entry, "faculty_2": ""}
                faculty_sol[f2.strip()][(d, p)] = entry2
    return dict(faculty_sol)

# ---------------------------------------------------------------------------
# Views
# ---------------------------------------------------------------------------


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def generate_timetable(request):
    """
    POST /api/timetable/generate/
    multipart/form-data:
      - Schedule config fields (see _build_config)
      - college_name  (optional)
      - file: Excel with branch/year/semester/section columns
    """
    try:
        config = _build_config(request.data)
    except ValueError as e:
        return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    if "file" not in request.FILES:
        return Response(
            {"error": "No Excel file uploaded."}, status=status.HTTP_400_BAD_REQUEST
        )

    try:
        subjects = _parse_excel(request.FILES["file"])
    except ValueError as e:
        return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    college_name = request.data.get("college_name", "SPMVV")
    groups = _derive_groups(subjects)

    # Parse optional custom_constraints JSON field
    import json as _json
    cc_raw = request.data.get("custom_constraints", None)
    if cc_raw:
        try:
            custom_constraints = _json.loads(cc_raw) if isinstance(cc_raw, str) else cc_raw
        except Exception:
            return Response(
                {"error": "custom_constraints must be valid JSON."},
                status=status.HTTP_400_BAD_REQUEST,
            )
    else:
        custom_constraints = {}

    # Inject sections into config for solver
    all_sections = sorted({sec for g in groups for sec in g["sections"]})
    config["sections"] = all_sections
    config["custom_constraints"] = custom_constraints

    try:
        solution = solve_timetable(config, subjects)
    except ValueError as solver_error:
        ai_explanation = _ai_explain_failure(str(solver_error), config)
        return Response(
            {"error": str(solver_error), "ai_explanation": ai_explanation},
            status=status.HTTP_422_UNPROCESSABLE_ENTITY,
        )

    ai_summary = _ai_summary(config, solution, groups, college_name)
    serializable = _solution_to_serializable(solution)

    faculty_solution = _build_faculty_solution(solution)
    faculty_serializable = _solution_to_serializable(faculty_solution)

    return Response(
        {
            "solution": serializable,
            "faculty_solution": faculty_serializable,
            "config": config,
            "groups": groups,
            "college_name": college_name,
            "ai_summary": ai_summary,
            "sections": list(solution.keys()),
            "faculty_list": sorted(faculty_solution.keys()),
        },
        status=status.HTTP_200_OK,
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def download_timetable(request):
    """
    POST /api/timetable/download/
    JSON body: { format, section, solution, config, groups, college_name }
    """
    body = request.data
    fmt = body.get("format", "pdf").lower()
    section = body.get("section", "")
    config = body.get("config", {})
    groups = body.get("groups", [])
    college_name = body.get("college_name", "SPMVV")
    sol_raw = body.get("solution", {})

    if not sol_raw or not config:
        return Response(
            {"error": "solution and config are required."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Build metadata for PDF/Excel generators (use first group as primary)
    if groups:
        g0 = groups[0]
        metadata = {
            "college_name": college_name,
            "branch": g0.get("branch", ""),
            "year": g0.get("year", ""),
            "semester": g0.get("semester", ""),
        }
    else:
        metadata = {
            "college_name": college_name,
            "branch": "",
            "year": "",
            "semester": "",
        }

    full_solution = _solution_from_serializable(sol_raw)

    if fmt == "pdf":
        if not section or section not in full_solution:
            return Response(
                {"error": "Section '{}' not found in solution.".format(section)},
                status=status.HTTP_400_BAD_REQUEST,
            )
        buf = generate_timetable_pdf(config, full_solution[section], section, metadata)
        fname = "timetable_{}_Y{}S{}_Sec{}.pdf".format(
            metadata.get("branch", "branch"),
            metadata.get("year", ""),
            metadata.get("semester", ""),
            section,
        )
        resp = HttpResponse(buf.read(), content_type="application/pdf")
        resp["Content-Disposition"] = 'attachment; filename="{}"'.format(fname)
        return resp

    elif fmt == "excel":
        buf = generate_timetable_excel(config, full_solution, metadata)
        fname = "timetable_{}_Y{}S{}.xlsx".format(
            metadata.get("branch", "branch"),
            metadata.get("year", ""),
            metadata.get("semester", ""),
        )
        resp = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="{}"'.format(fname)
        return resp

    return Response(
        {"error": "format must be pdf or excel."}, status=status.HTTP_400_BAD_REQUEST
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def timetable_sample_template(request):
    """GET /api/timetable/template/ -- returns 13-column sample Excel template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Subjects"

    required_headers = [
        "subject_code",
        "subject_name",
        "type",
        "hours_per_week",
        "faculty_1",
        "faculty_2",
        "branch",
        "year",
        "semester",
        "section",
    ]
    optional_headers = ["lab_consecutive_periods", "lab_split_mode", "min_faculty_gap"]
    all_headers = required_headers + optional_headers

    hdr_fill = PatternFill(fill_type="solid", fgColor="FF1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFFFF")
    opt_fill = PatternFill(fill_type="solid", fgColor="FFBDD7EE")
    opt_font = Font(bold=True, color="FF1E3A5F")

    ws.append(all_headers)
    for i, cell in enumerate(ws[1]):
        if i < len(required_headers):
            cell.fill = hdr_fill
            cell.font = hdr_font
        else:
            cell.fill = opt_fill
            cell.font = opt_font
        cell.alignment = Alignment(horizontal="center")

    sample_rows = [
        [
            "CS301",
            "Data Structures",
            "theory",
            4,
            "Dr. A Kumar",
            "",
            "CSE",
            "2",
            "1",
            "A",
            "",
            "",
            "",
        ],
        [
            "CS302",
            "Operating Systems",
            "theory",
            3,
            "Dr. B Rao",
            "",
            "CSE",
            "2",
            "1",
            "A",
            "",
            "",
            "",
        ],
        [
            "CS381",
            "DS Lab",
            "lab",
            3,
            "Dr. A Kumar",
            "Dr. C Reddy",
            "CSE",
            "2",
            "1",
            "A",
            3,
            "split",
            "",
        ],
        [
            "CS301",
            "Data Structures",
            "theory",
            4,
            "Dr. A Kumar",
            "",
            "CSE",
            "2",
            "1",
            "B",
            "",
            "",
            "",
        ],
        [
            "CS302",
            "Operating Systems",
            "theory",
            3,
            "Dr. B Rao",
            "",
            "CSE",
            "2",
            "1",
            "B",
            "",
            "",
            "",
        ],
        [
            "CS381",
            "DS Lab",
            "lab",
            3,
            "Dr. D Sharma",
            "Dr. E Rao",
            "CSE",
            "2",
            "1",
            "B",
            3,
            "split",
            "",
        ],
        [
            "EC301",
            "Signals & Systems",
            "theory",
            4,
            "Dr. F Nair",
            "",
            "ECE",
            "2",
            "1",
            "ALL",
            "",
            "",
            "",
        ],
        [
            "EC381",
            "Electronics Lab",
            "lab",
            2,
            "Dr. G Menon",
            "Dr. H Iyer",
            "ECE",
            "2",
            "1",
            "ALL",
            2,
            "whole",
            1,
        ],
    ]
    for row in sample_rows:
        ws.append(row)

    col_widths = [14, 24, 10, 14, 18, 18, 9, 6, 9, 9, 22, 16, 16]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ("subject_code", "required", "Unique subject identifier e.g. CS301"),
        ("subject_name", "required", "Full subject name"),
        ("type", "required", "theory or lab (lowercase)"),
        ("hours_per_week", "required", "Number of periods per week for this subject"),
        ("faculty_1", "required", "Primary faculty name"),
        (
            "faculty_2",
            "optional",
            "Second faculty for lab split (leave blank otherwise)",
        ),
        ("branch", "required", "Branch code e.g. CSE, ECE, MECH"),
        ("year", "required", "Academic year: 1, 2, 3 or 4"),
        ("semester", "required", "Semester: 1 or 2"),
        (
            "section",
            "required",
            "Section letter e.g. A, B -- or ALL to apply to every section",
        ),
        (
            "lab_consecutive_periods",
            "optional",
            "Override global: consecutive periods for this lab session",
        ),
        (
            "lab_split_mode",
            "optional",
            "split (halves run in parallel) or whole (entire section)",
        ),
        (
            "min_faculty_gap",
            "optional",
            "Override global: min free periods between classes for this faculty",
        ),
    ]
    ws2.append(["Column", "Required?", "Description"])
    for cell in ws2[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
    for row in instructions:
        ws2.append(list(row))
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 62

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="timetable_template.xlsx"'
    return resp


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def download_all_timetable(request):
    """
    POST /api/timetable/download-all/
    JSON body: { solution, config, groups, college_name }
    Returns a combined PDF with one page per section.
    """
    body         = request.data
    config       = body.get("config", {})
    groups       = body.get("groups", [])
    college_name = body.get("college_name", "SPMVV")
    sol_raw      = body.get("solution", {})

    if not sol_raw or not config:
        return Response(
            {"error": "solution and config are required."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    full_solution = _solution_from_serializable(sol_raw)
    buf = generate_all_timetables_pdf(config, full_solution, groups, college_name)
    resp = HttpResponse(buf.read(), content_type="application/pdf")
    resp["Content-Disposition"] = 'attachment; filename="timetable_all_sections.pdf"'
    return resp


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def download_faculty_timetable(request):
    """
    POST /api/timetable/download-faculty/
    JSON body: { faculty_name, faculty_solution, config, college_name }
    Returns PDF for a single faculty member.
    """
    body         = request.data
    faculty_name = body.get("faculty_name", "")
    config       = body.get("config", {})
    college_name = body.get("college_name", "SPMVV")
    sol_raw      = body.get("faculty_solution", {})

    if not faculty_name:
        return Response({"error": "faculty_name is required."}, status=status.HTTP_400_BAD_REQUEST)
    if not sol_raw or not config:
        return Response({"error": "faculty_solution and config are required."}, status=status.HTTP_400_BAD_REQUEST)

    if faculty_name not in sol_raw:
        return Response(
            {"error": "Faculty '{}' not found in faculty_solution.".format(faculty_name)},
            status=status.HTTP_400_BAD_REQUEST,
        )

    fac_sol_raw = sol_raw[faculty_name]
    fac_sol = {}
    for key_str, info in fac_sol_raw.items():
        d, p = map(int, key_str.split(","))
        fac_sol[(d, p)] = info

    buf = generate_faculty_timetable_pdf(config, fac_sol, faculty_name, college_name)
    safe_name = faculty_name.replace(" ", "_").replace(".", "")
    fname = "timetable_faculty_{}.pdf".format(safe_name)
    resp = HttpResponse(buf.read(), content_type="application/pdf")
    resp["Content-Disposition"] = 'attachment; filename="{}"'.format(fname)
    return resp


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def download_all_faculty_timetables(request):
    """
    POST /api/timetable/download-faculty-all/
    JSON body: { faculty_solution, config, college_name }
    Returns combined PDF with one page per faculty.
    """
    body         = request.data
    config       = body.get("config", {})
    college_name = body.get("college_name", "SPMVV")
    sol_raw      = body.get("faculty_solution", {})

    if not sol_raw or not config:
        return Response({"error": "faculty_solution and config are required."}, status=status.HTTP_400_BAD_REQUEST)

    full_faculty_solution = {}
    for faculty_name, slot_map in sol_raw.items():
        fac_sol = {}
        for key_str, info in slot_map.items():
            d, p = map(int, key_str.split(","))
            fac_sol[(d, p)] = info
        full_faculty_solution[faculty_name] = fac_sol

    buf = generate_all_faculty_timetables_pdf(config, full_faculty_solution, college_name)
    resp = HttpResponse(buf.read(), content_type="application/pdf")
    resp["Content-Disposition"] = 'attachment; filename="timetable_all_faculty.pdf"'
    return resp
