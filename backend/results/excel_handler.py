import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from django.conf import settings
import logging

logger = logging.getLogger('django')

class ExcelValidator:
    REQUIRED_COLUMNS = [
        'Branch', 'Roll Number', 'Student Name', 'Result Type'
    ]
    
    ALLOWED_EXTENSIONS = ['.xlsx', '.xls']
    MAX_FILE_SIZE = settings.FILE_UPLOAD_MAX_MEMORY_SIZE
    
    @staticmethod
    def validate_file(file):
        """Validate file type and size"""
        errors = []
        
        # Check file extension
        if not any(file.name.lower().endswith(ext) for ext in ExcelValidator.ALLOWED_EXTENSIONS):
            errors.append(f"Invalid file format. Allowed formats: {', '.join(ExcelValidator.ALLOWED_EXTENSIONS)}")
        
        # Check file size
        if file.size > ExcelValidator.MAX_FILE_SIZE:
            errors.append(f"File size exceeds maximum allowed size of {ExcelValidator.MAX_FILE_SIZE / (1024*1024)}MB")
        
        return errors
    
    @staticmethod
    def validate_structure(workbook):
        """Validate Excel structure"""
        errors = []
        
        try:
            sheet = workbook.active
            
            if sheet.max_row < 2:
                errors.append("Excel file is empty or has no data rows")
                return errors
            
            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
            
            # Check required columns
            for required_col in ExcelValidator.REQUIRED_COLUMNS:
                if required_col not in headers:
                    errors.append(f"Missing required column: {required_col}")
            
        except Exception as e:
            errors.append(f"Error reading Excel structure: {str(e)}")
        
        return errors
    
    @staticmethod
    def normalize_result(result_value):
        """Normalize result value to proper case (Pass/Fail/Absent)"""
        if not result_value or not isinstance(result_value, str):
            return result_value
        
        normalized = result_value.strip().lower()
        if normalized == 'pass':
            return 'Pass'
        elif normalized == 'fail':
            return 'Fail'
        elif normalized == 'absent':
            return 'Absent'
        return result_value
    
    @staticmethod
    def normalize_grade(grade_value):
        """Normalize grade value to uppercase"""
        if not grade_value or not isinstance(grade_value, str):
            return grade_value
        return grade_value.strip().upper()
    
    @staticmethod
    def normalize_branch(branch_value):
        """Normalize branch value to lowercase"""
        if not branch_value or not isinstance(branch_value, str):
            return 'cse'  # Default branch
        
        # Map common branch names to standard codes
        branch_map = {
            'computer science': 'cse',
            'computer science and engineering': 'cse',
            'cse': 'cse',
            'cs': 'cse',
            'electronics and communication': 'ece',
            'electronics and communication engineering': 'ece',
            'ece': 'ece',
            'ec': 'ece',
            'electrical and electronics': 'eee',
            'electrical and electronics engineering': 'eee',
            'eee': 'eee',
            'ee': 'eee',
            'mechanical': 'mech',
            'mechanical engineering': 'mech',
            'mech': 'mech',
            'me': 'mech',  # ME is commonly used for Mechanical
            'civil': 'civil',
            'civil engineering': 'civil',
            'ce': 'civil',
            'information technology': 'it',
            'it': 'it',
            'chemical': 'chemical',
            'chemical engineering': 'chemical',
            'che': 'chemical',
            'biotechnology': 'biotechnology',
            'biotech': 'biotechnology',
            'bt': 'biotechnology',
            'mba': 'mba',
            'mca': 'mca'
        }
        
        normalized = branch_value.strip().lower()
        return branch_map.get(normalized, normalized)
    
    @staticmethod
    def normalize_result_type(result_type_value):
        """Normalize result type value"""
        if not result_type_value or not isinstance(result_type_value, str):
            return 'regular'  # Default
        
        # Map common variations
        result_type_map = {
            'regular': 'regular',
            'supplementary': 'supplementary',
            'supply': 'supplementary',
            'supple': 'supplementary',
            'both': 'both',
            'regular and supplementary': 'both',
            'regular & supplementary': 'both',
            'all': 'both'
        }
        
        normalized = result_type_value.strip().lower()
        return result_type_map.get(normalized, 'regular')
    
    @staticmethod
    def parse_excel(file):
        """Parse Excel file and return structured data"""
        try:
            workbook = openpyxl.load_workbook(file, data_only=True)
            sheet = workbook.active
            
            # Get headers
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
            
            # Parse data rows
            data = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Skip empty rows
                    continue
                
                row_data = {}
                subjects = []
                
                for idx, value in enumerate(row):
                    if idx >= len(headers):
                        break
                    
                    header = headers[idx]
                    
                    # Branch field
                    if header == 'Branch':
                        row_data[header] = ExcelValidator.normalize_branch(value)
                    # Result Type field (NEW)
                    elif header == 'Result Type':
                        row_data[header] = ExcelValidator.normalize_result_type(value)
                    # Basic fields
                    elif header in ['Roll Number', 'Student Name']:
                        row_data[header] = value
                    elif header == 'Total Marks':
                        # Handle total marks value
                        row_data[header] = int(value) if value is not None and value != '' else None
                    elif header == 'SGPA':
                        # Handle SGPA value
                        row_data[header] = float(value) if value is not None and value != '' else None
                    # Subject fields
                    elif 'Subject' in header:
                        parts = header.split()
                        if len(parts) >= 3:
                            subject_num = parts[1]
                            field_name = ' '.join(parts[2:])
                            
                            # Find or create subject entry
                            subject = next((s for s in subjects if s.get('number') == subject_num), None)
                            if not subject:
                                subject = {'number': subject_num}
                                subjects.append(subject)
                            
                            # Normalize Grade field for subjects
                            if field_name == 'Grade':
                                subject[field_name] = ExcelValidator.normalize_grade(value)
                            # Handle Credits field
                            elif field_name == 'Credits':
                                subject[field_name] = int(value) if value is not None and value != '' else None
                            else:
                                subject[field_name] = value
                
                row_data['subjects'] = subjects
                row_data['row_number'] = row_idx
                data.append(row_data)
            
            return data, None
            
        except InvalidFileException:
            return None, ["Invalid Excel file format"]
        except Exception as e:
            logger.error(f"Error parsing Excel: {str(e)}")
            return None, [f"Error parsing Excel file: {str(e)}"]
    
    @staticmethod
    def validate_data(data):
        """Validate parsed data (case-insensitive validation)"""
        errors = []
        
        valid_result_types = ['regular', 'supplementary', 'both']
        valid_grades = ['O', 'A', 'B', 'C', 'D', 'F']  # Uppercase for comparison, simplified
        valid_branches = ['cse', 'ece', 'eee', 'mech', 'civil', 'it', 'chemical', 'biotechnology', 'mba', 'mca']
        
        for row in data:
            row_num = row.get('row_number', 'Unknown')
            
            # Validate Branch
            branch = row.get('Branch', '')
            if not branch:
                errors.append(f"Row {row_num}: Branch is required")
            elif branch not in valid_branches:
                errors.append(f"Row {row_num}: Invalid Branch '{branch}'. Must be one of: {', '.join([b.upper() for b in valid_branches])}")
            
            # Validate Result Type (NEW)
            result_type = row.get('Result Type', '')
            if not result_type:
                errors.append(f"Row {row_num}: Result Type is required")
            elif result_type not in valid_result_types:
                errors.append(f"Row {row_num}: Invalid Result Type '{result_type}'. Must be Regular, Supplementary, or Both")
            
            # Validate required fields
            if not row.get('Roll Number'):
                errors.append(f"Row {row_num}: Roll Number is required")
            
            if not row.get('Student Name'):
                errors.append(f"Row {row_num}: Student Name is required")
            
            # Validate Total Marks
            total_marks = row.get('Total Marks', None)
            if total_marks is not None and not isinstance(total_marks, (int, float)):
                errors.append(f"Row {row_num}: Total Marks must be a number")
            
            # Validate SGPA
            sgpa = row.get('SGPA', None)
            if sgpa is not None and not isinstance(sgpa, (int, float)):
                errors.append(f"Row {row_num}: SGPA must be a number")
            elif sgpa is not None and (sgpa < 0 or sgpa > 10):
                errors.append(f"Row {row_num}: SGPA must be between 0 and 10")
            
            # Validate subjects
            subjects = row.get('subjects', [])
            if not subjects:
                errors.append(f"Row {row_num}: No subjects found")
            
            for subject in subjects:
                subject_num = subject.get('number', 'Unknown')
                
                if not subject.get('Code'):
                    errors.append(f"Row {row_num}, Subject {subject_num}: Subject Code is required")
                
                if not subject.get('Name'):
                    errors.append(f"Row {row_num}, Subject {subject_num}: Subject Name is required")
                
                # Validate credits
                credits = subject.get('Credits', None)
                if credits is None:
                    errors.append(f"Row {row_num}, Subject {subject_num}: Credits is required")
                elif not isinstance(credits, int) or credits < 0:
                    errors.append(f"Row {row_num}, Subject {subject_num}: Credits must be a positive integer")
                
                # Validate subject grade (case-insensitive)
                subject_grade = subject.get('Grade', '')
                if not subject_grade:
                    errors.append(f"Row {row_num}, Subject {subject_num}: Grade is required")
                elif subject_grade.strip().upper() not in valid_grades:
                    errors.append(f"Row {row_num}, Subject {subject_num}: Invalid Grade. Must be one of: {', '.join([g.upper() for g in valid_grades])} (case-insensitive)")
        
        return errors
